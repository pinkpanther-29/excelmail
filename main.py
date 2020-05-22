import openpyxl
import os
import pandas as pd
os.chdir('/Users/TalishaGriesbach/Desktop')

file = openpyxl.load_workbook('mailinglist.xlsx')['sheet']
cell = file['A5']
print(type(cell.value)) # already a string

for i in range (2,35):
    print(i, file.cell(row=i, column=1).value, file.cell(row=i, column=2).value)